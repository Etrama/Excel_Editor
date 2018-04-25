# -*- coding: utf-8 -*-
"""
Created on Sun Feb 11 19:49:35 2018

@author: kmoudgalya
"""

"""
This is an end to end excel scrubber script.
It takes of the following:
    1.Removing the version tab from excel files.
    2.Editing Specific cells in excel.
    3.Replacing a particular string across all sheets in a document.
    4.Updating Headers and Footers.
    5.Changing the file's prefix.
    
NOTE:
    THE SCRIPT PROBABLY DOES NOT WORK WHEN THERE ARE IMAGES IN EXCEL FILES.
    THIS IS NOT THE ONLY REASON FOR FAILURE THOUGH.
    THE SCRIPT ALSO DOES NOT HANDLE COMMENTS IN EXCEL.
"""
import openpyxl as xl
import os
from shutil import move 
from django.utils.encoding import smart_str
import time
#ppts are not in scope for this script!!!
location = raw_input(r"Enter the directory path where the files to be scrubbed are located: ")
# progresbar installed with:
#conda install -c anaconda progressbar 
#use this in windows prompt, without python open.
filenames = os.listdir(location)

def version_tab_deleter(location,filenames):
    for fname in filenames:
        if os.path.splitext(fname)[1] == '.xlsx':
            print "Removing version tab from: " + fname
            workbook = xl.load_workbook(os.path.join(location,fname))
            std = workbook.get_sheet_by_name('Version')
            workbook.remove_sheet(std)
            workbook.save(os.path.join(location,fname))
            print "Version Tab deleted successfully."
    print('\007')
    print('\007')
    print('\007')

#If your spreadsheet is very large, you can add\
#an argument on_demand to open_workbook, which loads only current sheets to memory
        
def specific_cell_editor(location,filenames):
    sheet_name = raw_input("Enter the sheet name where the cell to be edited is located, such 'Sheet 1': ")
    cell_location = raw_input("Enter the location of the cell which is to be edited: ")
    for fname in filenames:
         if os.path.splitext(fname)[1] == '.xlsx':
            print "File being edited: " + fname
    
            wb = xl.load_workbook(os.path.join(location,fname))
            ws = wb[sheet_name]
            ws[cell_location] = ""
            wb.save(os.path.join(location,fname))
            print "Cell " + cell_location + " edited successfully."
    print('\007')
    print('\007')
    print('\007')

 
"""
Pickle let's you use data generated fro mone script in another.
#fileObject = open(os.path.join("C:\\python projects\\","intermediary1.pkl"))
#failedFileList = pickle.load(fileObject)    """ 
  
def general_cell_editor(location,filenames):
    old_string = raw_input("Enter string to be replaced: ")
    new_string = raw_input("Enter new string which would take its place: ")
    
    failedGeneralCellEditor = []
    for fname in filenames:
         if os.path.splitext(fname)[1] == '.xlsx':
            try:
                nextfileflag = 0
                print "File being edited: " + fname 
                wb1 = xl.load_workbook(os.path.join(location,fname))
                ws1 = wb1.worksheets[0]
                time_1 = time.time()
                for ws1 in wb1.worksheets:
                    if nextfileflag == 1:
                            break
                    for row in ws1:
                        if nextfileflag == 1:
                            break
                        for cell in row:
                            temp = smart_str(cell.value)
                            if temp != "None":
                                ws1[cell.coordinate].value = temp.replace(old_string,new_string)
                                #time_2 = time.time()
                                #print time_2 - time_1
                                if time_2 - time_1 >= 30.00:
                                    failedGeneralCellEditor.append(fname)
                                    nextfileflag = 1
                                    break
                                
                if nextfileflag == 1:
                    print "Skipped, the string " + old_string + " was not completely replaced in this file."
                    print ""
                    wb1.save(os.path.join(location,fname))
                    continue                                
                wb1.save(os.path.join(location,fname))
                print "Replaced " + old_string + " with " + new_string + " successfully."
                print ""
                
            except:
                print "Skipped, file was NOT edited."
                print ""
                failedGeneralCellEditor.append(fname)
                continue
    
    if len(failedGeneralCellEditor) != 0 :     
        for name in failedGeneralCellEditor:
            if not os.path.exists(os.path.join(location,name)):
                print "Do manually scrub the client references in the following files: "
                print failedGeneralCellEditor
                print ""
                print "These files will be shifted to the folder 'Files which need manual cell string replacement' "
                if os.path.exists(os.path.join(location,"Files which need manual cell string replacement")):
                    for fname in failedGeneralCellEditor:
                        move(os.path.join(location,fname),os.path.join(location,"Files which need manual cell string replacement"))
                else:
                    os.makedirs(os.path.join(location,"Files which need manual cell string replacement"))
                    for fname in failedGeneralCellEditor:
                        move(os.path.join(location,fname),os.path.join(location,"Files which need manual cell string replacement"))
            else:
                print "Rejoice! None of the files need manual client reference scrubbing."
 
    else:
        print "Rejoice! None of the files need manual client reference scrubbing."

    print('\007')
    print('\007')
    print('\007')


"""#fileObject = open(os.path.join("C:\\python projects\\","intermediary1.pkl"),'w')
#pickle.dump(failedHeaderFooterUpdate,fileObject)"""

def header_footer_updater(location,filenames):
    failedHeaderFooterUpdate = []
    
    for fname in filenames:
        if os.path.splitext(fname)[1] == '.xlsx':
            try:
                nextfileflag = 0
                time_1 = time.time()
                print "File being edited: " + fname 
                
                wb1 = xl.load_workbook(os.path.join(location,fname))
                ws1 = wb1.worksheets[0]
                
                for ws1 in wb1.worksheets:
                    if nextfileflag == 1:
                        break
                    ws1.oddHeader.center.text = ws1.title
                    ws1.oddHeader.center.size = 12
                    ws1.oddHeader.center.font = "Arial"
                    ws1.oddFooter.center.text = "This document is intended solely for the information and internal use of XX and should not be used or relied upon by any other person or entity."
                    ws1.oddFooter.center.size = 12
                    ws1.oddFooter.center.font = "Arial"
                    
                    time_2 = time.time()
                    if time_2 - time_1 >= 75.00:
                        failedHeaderFooterUpdate.append(fname)
                        nextfileflag = 1
                        break
                if nextfileflag == 1:
                    print "Skipped, Headers and Footers were NOT updated."
                    print ""
                    continue
                
                wb1.save(os.path.join(location,fname))
                print "Headers and Footers were updated successfully."
                print ""
            except:
                print "Skipped, Headers and Footers were NOT updated."
                print ""
                failedHeaderFooterUpdate.append(fname)
                continue
            
    if len(failedHeaderFooterUpdate) != 0:            
        for name in failedHeaderFooterUpdate:
            if not os.path.exists(os.path.join(location,name)):
                print "Manually update the Headers and Footers for the following files: "
                print failedHeaderFooterUpdate
                print "These files will be shifted to the folder 'Files which need manual Header and Footer updates' "
                if os.path.exists(os.path.join(location,"Files which need manual Header and Footer updates")):
                    for fname in failedHeaderFooterUpdate:
                        move(os.path.join(location,fname),os.path.join(os.path.join(location,fname),"Files which need manual Header and Footer updates"))
                else:
                    os.makedirs(os.path.join(location,"Files which need manual Header and Footer updates"))
                    for fname in failedHeaderFooterUpdate:
                        move(os.path.join(location,fname),os.path.join(location,"Files which need manual Header and Footer updates"))
            else:
                print "Rejoice! None of the files need manual updation of headers and footers."
 
    else:
        print "Rejoice! None of the files need manual updation of headers and footers."
    print('\007')
    print('\007')
    print('\007')


def change_prefix(location,filenames):
    prefix = raw_input("Enter the common prefix the files have: ")
    newprefix = raw_input("Enter the new prefix for these files: ")
    
    filenames = os.listdir(location)
    
    for fname in filenames:
         if os.path.splitext(fname)[1] == '.xlsx':
            if fname.startswith(prefix):
                os.rename(os.path.join(location,fname),os.path.join(location,fname.replace(prefix,newprefix)))
    print('\007')
    print('\007')
    print('\007')

