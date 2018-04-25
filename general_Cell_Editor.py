# -*- coding: utf-8 -*-
"""
Created on Tue Feb 06 17:00:04 2018

@author: kmoudgalya
"""
import os
import openpyxl as xl
#import pickle
from shutil import move
from django.utils.encoding import smart_str

#print "Please note the location variable must be the same across header_Footer_Updater.py\
#and general_Cell_editor.py"
print ""
#print "This script assumes that a 'Files which errored out' folder was created\
#in header_Footer_updater.py"

location = raw_input(r"Enter the directory path where the files whose strings\
 need editing are located: ")
filenames = os.listdir(location)


old_string = raw_input("Enter string to be replaced: ")
new_string = raw_input("Enter new string which would take its place: ")

#fileObject = open(os.path.join("C:\\python projects\\","intermediary1.pkl"))
#failedFileList = pickle.load(fileObject)
failedCellEditor = []
for fname in filenames:
    try:
        print "File being edited: " + fname
        
        wb1 = xl.load_workbook(os.path.join(location, fname))
        ws1 = wb1.worksheets[0]
        
        wb2 = xl.Workbook()
        
        for ws1 in wb1.worksheets:
            ws2 = wb2.create_sheet(ws1.title)
            for row in ws1:
                for cell in row:
                    temp = smart_str(cell.value)
                    if temp != "None":
                        ws2[cell.coordinate].value = temp.replace(old_string, new_string)

        
        std = wb2.get_sheet_by_name('Sheet')
        wb2.remove_sheet(std)        
        wb2.save(os.path.join(location,"GE"+fname))
        print "New file created successfully."
        print ""
        
    except:
        print "Skipped, file was NOT edited."
        print ""
        failedCellEditor.append(fname)
        continue
        
print "Do manually scrub the client references in the following files: "
print failedCellEditor
print ""

#os.makedirs(os.path.join("C:\python projects\Files which need manual scrubbing","Files which need manual cell string replacement"))
print "These files will be shifted to the folder 'Files which need manual cell string replacement' inside the folder 'Files which need manual scrubbing'."

#for fname in failedCellEditor:
#    move(os.path.join(location,fname),"C:\python projects\Files which need manual scrubbing\Files which need manual cell string replacement")
