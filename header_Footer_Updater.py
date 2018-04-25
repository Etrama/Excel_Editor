# -*- coding: utf-8 -*-
"""
Created on Sun Feb 04 19:22:14 2018

@author: kmoudgalya
"""

#conda install -c anaconda xlsxwriter 
#use the above to install xlsxwriter
import os
import openpyxl as xl
import time
#import pickle 
from shutil import move

#shutil copy of one excel file to another copies the headers and footers as well.
location = raw_input(r"Enter the directory path where the files whose headers \
                     and footers need editing are located: ")
filenames = os.listdir(location)

failedHeaderFooterUpdate = []

#code to copy one excel to another without headers or footers from the source file
for fname in filenames:
    try:
        nextfileflag = 0
        time_1 = time.time()
        print "File being edited: " + fname 
        
        wb1 = xl.load_workbook(os.path.join(location,fname))
        ws1 = wb1.worksheets[0]
        
        wb2 = xl.Workbook()
        
        for ws1 in wb1.worksheets:
            if nextfileflag == 1:
                break
            ws2 = wb2.create_sheet(ws1.title)
            ws2.oddHeader.center.text = ws1.title
            ws2.oddHeader.center.size = 12
            ws2.oddHeader.center.font = "Arial"
            ws2.oddFooter.center.text = "This document is intended solely for the information and internal use of XX and should not be used or relied upon by any other person or entity."
            ws2.oddFooter.center.size = 12
            ws2.oddFooter.center.font = "Arial"
            for row in ws1:
                if nextfileflag == 1:
                    break
                for cell in row:
                    ws2[cell.coordinate].value = cell.value
                    time_2 = time.time()
                    if time_2 - time_1 >= 30.00:
                        failedHeaderFooterUpdate.append(fname)
                        nextfileflag = 1
                        break
        if nextfileflag == 1:
            print "Skipped, file was NOT created."
            print ""
            continue
        
        std = wb2.get_sheet_by_name('Sheet')
        wb2.remove_sheet(std)        
        wb2.save(os.path.join(location,"HF"+fname))
        print "New file created successfully."
        print ""
    except:
        print "Skipped, file was NOT created."
        print ""
        failedHeaderFooterUpdate.append(fname)
        continue
        
        
#fileObject = open(os.path.join("C:\\python projects\\","intermediary1.pkl"),'w')
#pickle.dump(failedHeaderFooterUpdate,fileObject)
print "Do manually scrub the following files: "
print failedHeaderFooterUpdate

location2 = raw_input("These files will be shifted to a new folder. \
                      Enter the location where you want the folder to be created: " )
os.makedirs(os.path.join(location2,"Files which need manual scrubbing"))
for fname in failedHeaderFooterUpdate:
    move(os.path.join(location,fname),os.path.join(location2,"Files which need manual scrubbing"))
    
